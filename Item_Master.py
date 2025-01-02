import os
import shutil
import pandas as pd
from datetime import datetime
import openai
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Define paths
IMPORT_PATH = r"C:\Users\it\OneDrive\Documents\_Item Masters\Import"
EXPORT_PATH = r"C:\Users\it\OneDrive\Documents\_Item Masters\Export"
TEMPLATE_PATH = r"C:\Users\it\OneDrive\Documents\_Item Masters\_Template\template.xlsx"

IMPORT_ARCHIVE_PATH = os.path.join(IMPORT_PATH, "Archive")
EXPORT_ARCHIVE_PATH = os.path.join(EXPORT_PATH, "Archive")

# Ensure archive directories exist
os.makedirs(IMPORT_ARCHIVE_PATH, exist_ok=True)
os.makedirs(EXPORT_ARCHIVE_PATH, exist_ok=True)

# Set up OpenAI API key
openai.api_key = "sk-jlfEUEj_tr1lpYN3W5ZJcV4Ei1kLy8o4SGQo1HBZuAT3BlbkFJexbZEnhwyFFroFZmVAzM0SeGLWNx5xTXo35SySZSwA"


def archive_export_files():
    """Move existing files in the Export directory to the Export Archive."""
    for filename in os.listdir(EXPORT_PATH):
        file_path = os.path.join(EXPORT_PATH, filename)
        if os.path.isfile(file_path):
            shutil.move(file_path, os.path.join(EXPORT_ARCHIVE_PATH, filename))
            print(f"Archived export file: {filename}")


def generate_suggested_description(description):
    """Generate a suggested description using GPT-4 API if it's over 50 characters."""
    prompt = (
        f"Please shorten the following product description to 50 characters or less, ensuring that the meaning remains clear. "
        f"Do not exceed 50 characters under any circumstances. Description: '{description}'"
    )

    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[
                {"role": "system",
                 "content": "You are an assistant that helps shorten product descriptions to be concise yet meaningful, strictly adhering to a 50-character limit."},
                {"role": "user", "content": prompt},
            ]
        )
        shortened_description = response['choices'][0]['message']['content'].strip()

        # Ensure the response is exactly 50 characters or less
        if len(shortened_description) > 50:
            shortened_description = shortened_description[:47] + "..."

        return shortened_description
    except Exception as e:
        print(f"Error generating suggested description: {e}")
        return None


def remove_tables(ws):
    """Remove all tables from the worksheet."""
    for table_name in list(ws.tables.keys()):
        del ws.tables[table_name]


def process_import_files():
    """Process all files in the Import directory."""
    import_files = [f for f in os.listdir(IMPORT_PATH) if
                    os.path.isfile(os.path.join(IMPORT_PATH, f)) and f.lower().endswith(('.xlsx', '.xls'))]

    if not import_files:
        print("No import files found to process.")
        return

    for file in import_files:
        file_path = os.path.join(IMPORT_PATH, file)
        client_name = file.split('_')[0]  # Extract client name using only the first underscore
        process_date = datetime.now().strftime('%m-%d-%y')
        needs_review = False

        try:
            print(f"\nProcessing file: {file}")

            # Load data from import file
            import_df = pd.read_excel(file_path, header=2)  # Assuming headers are on the 3rd row
            print("Import file loaded successfully.")

            # Load template workbook and select the active sheet
            wb_template = load_workbook(TEMPLATE_PATH)
            ws_template = wb_template.active
            print("Template file loaded successfully.")

            # Identify the table in the template
            table = None
            if ws_template.tables:
                table = ws_template.tables[next(iter(ws_template.tables))]
                print(f"Table '{table.name}' found in template.")

            # Remove the table from the template sheet
            remove_tables(ws_template)

            # Prepare a DataFrame to hold the processed data
            template_headers = [cell.value for cell in ws_template[3]]
            processed_data = pd.DataFrame(columns=template_headers)

            # Print import data columns
            print("Import Data Columns:", import_df.columns.tolist())

            # Map columns to be transferred
            columns_to_transfer = [
                "Description", "Description 2", "Created From Nonstock Item", "Item No", "Category", "Pack type",
                "Count unit", "Alt Unit 1", "Alt unit 1 tracked", "Alt unit 2", "Alt unit 2 tracked",
                "Weight unit", "Dimension unit", "Convert 0 to 1 factor", "Check Unique Tag ID",
                "Sub 1 Type", "1st Pick Sort By", "2nd Pick Sort By", "Guided Pick", "Use Pick Replenishment",
                "UPC Code"
            ]

            # Filter rows where 'Item No' is not empty
            import_df = import_df[import_df['Item No'].notna()]

            # Replace '&' with '/' in descriptions
            import_df['Description'] = import_df['Description'].str.replace('&', '/', regex=False)

            # Analyze descriptions and prepare for API processing
            long_descriptions = import_df['Description'].apply(lambda x: len(x) > 50)
            import_df["Suggested Description"] = None

            # Process only long descriptions with the API
            if long_descriptions.any():
                for idx in import_df[long_descriptions].index:
                    desc = import_df.at[idx, "Description"]
                    import_df.at[idx, "Suggested Description"] = generate_suggested_description(desc)
                    needs_review = True

            # Assign predefined values
            predefined_values = {
                "Pack Type Updates": "Always Allow Pack Type Updates",
                "Category Updates": "Always Allow Category Updates",
                "Guided Pick": "None",
                "Use Pick Replenishment": "No",
                "Check Unique Tag ID": "Do Not Check",
                "Sub 1 Type": "Not used",
                "1st Pick Sort By": "Lowest Quantity",
                "2nd Pick Sort By": "SF Pick Rule",
                "Convert 0 to 1 meth": "Divide by",
                "Convert 1 to 2 meth": "Divide by",
                "Convert 0 to 2 meth": "Divide by",
                "Client": client_name
            }

            # Populate the processed DataFrame with data
            for col in template_headers:
                if col in import_df.columns and col in columns_to_transfer:
                    processed_data[col] = import_df[col]
                elif col in predefined_values:
                    processed_data[col] = predefined_values[col]
                else:
                    processed_data[col] = None  # Fill with NaN or appropriate default

            # Ensure all column headers are strings
            processed_data.columns = [str(col) for col in processed_data.columns]

            # Add "Suggested Description" to processed_data only for long descriptions
            if "Suggested Description" in import_df.columns and not import_df["Suggested Description"].isna().all():
                processed_data["Suggested Description"] = import_df["Suggested Description"]
            else:
                processed_data.drop(columns=["Suggested Description"], inplace=True)

            # Check if the number of columns matches
            if len(processed_data.columns) != len(template_headers) + (
            1 if "Suggested Description" in processed_data.columns else 0):
                raise ValueError("Mismatch between template headers and processed data columns")

            # Find the first empty row after the template headers
            start_row = 4  # Data should start from row 4
            for row in ws_template.iter_rows(min_row=4, max_row=ws_template.max_row):
                if not any(cell.value for cell in row):
                    break
                start_row += 1

            # Append the processed data to the template starting at the first empty row
            for r_idx, row in enumerate(dataframe_to_rows(processed_data, index=False, header=False), start=start_row):
                for c_idx, value in enumerate(row, 1):
                    ws_template.cell(row=r_idx, column=c_idx, value=value)

            # Calculate the correct range for the table including headers
            last_column = len(processed_data.columns)
            last_column_letter = ws_template.cell(row=3, column=last_column).column_letter
            end_row = start_row - 1 + len(processed_data)
            new_table_range = f"A3:{last_column_letter}{end_row}"
            new_table = Table(displayName=table.name, ref=new_table_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True,
            )
            new_table.tableStyleInfo = style
            ws_template.add_table(new_table)
            print(f"Table '{table.name}' recreated in the export file with range {new_table_range}.")

            # Generate export filename
            filename_suffix = f"{client_name} - {process_date} - PROCESSED"
            if needs_review:
                filename_suffix += " (REVIEW)"
            export_filename = f"{filename_suffix}.xlsx"
            export_filepath = os.path.join(EXPORT_PATH, export_filename)

            # Save the workbook with the appended data
            wb_template.save(export_filepath)
            print(f"Exported processed file: {export_filename}")

            # Move processed import file to archive
            shutil.move(file_path, os.path.join(IMPORT_ARCHIVE_PATH, file))
            print(f"Archived import file: {file}")

        except Exception as e:
            print(f"An error occurred while processing {file}: {e}")


if __name__ == "__main__":
    print("Starting data processing...")
    archive_export_files()
    process_import_files()
    print("Data processing completed.")
